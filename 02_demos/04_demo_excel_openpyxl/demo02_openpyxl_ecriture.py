from openpyxl import Workbook

print("=== Démo OpenPyXL - Écriture basique ===")

# Créer un nouveau classeur Excel en mémoire
wb = Workbook()

# Récupérer la feuille active
ws = wb.active

# Renommer la feuille
ws.title = "Ventes"

# Écrire les en-têtes dans les cellules A1, B1, C1
ws['A1'] = 'Produit'
ws['B1'] = 'Quantité'
ws['C1'] = 'Prix'

# Écrire une ligne de données avec cell(row, column, value)
ws.cell(row=2, column=1, value='Laptop')
ws.cell(row=2, column=2, value=5)
ws.cell(row=2, column=3, value=899.99)

# Ajouter plusieurs lignes d'un coup avec ws.append()
data = [
    ['Souris', 10, 29.99],
    ['Clavier', 3, 79.99],
    ['Écran', 2, 299.99]
]
for row in data:
    ws.append(row)

print("Contenu écrit dans la feuille 'Ventes' :")
for row in ws.iter_rows(values_only=True):
    print(row)

# Sauvegarder le classeur dans un fichier Excel
wb.save('output.xlsx')
print("\nFichier 'output.xlsx' créé.")
