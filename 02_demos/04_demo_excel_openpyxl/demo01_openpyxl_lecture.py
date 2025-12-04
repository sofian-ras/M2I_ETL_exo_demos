from openpyxl import load_workbook

print("=== Démo OpenPyXL - Lecture basique ===")

# Charger un fichier Excel existant
wb = load_workbook('data.xlsx')

# Lister les feuilles disponibles dans le classeur
print("\nFeuilles disponibles dans le classeur :")
print(wb.sheetnames)  # ['Feuil1', 'Feuil2', 'Ventes', ...]

# Sélectionner une feuille
ws = wb.active  # Feuille active (en général la première)
print(f"\nFeuille active : {ws.title}")

# Sélectionner une feuille par son nom
ws = wb['Feuil1']
print(f"Feuille sélectionnée par nom : {ws.title}")

# Lire une cellule (par coordonnées Excel)
valeur = ws['A1'].value  # Cellule A1
print(f"\nValeur de A1 (notation 'A1') : {valeur}")

# Lire une cellule (par indices de ligne/colonne, à partir de 1)
valeur = ws.cell(row=1, column=1).value  # Ligne 1, Colonne 1
print(f"Valeur de (row=1, col=1) : {valeur}")

# Lire une plage de cellules (A1 à C5)
print("\nLecture de la plage A1:C5 :")
for row in ws['A1:C5']:
    for cell in row:
        # Affiche la valeur de chaque cellule sur la même ligne, séparée par des tabulations
        print(cell.value, end='\t')
    print()  # Retour à la ligne après chaque ligne de la plage

# Itérer sur toutes les lignes (avec values_only=True pour obtenir directement les valeurs)
print("\nItération sur toutes les lignes de la feuille (values_only=True) :")
for row in ws.iter_rows(min_row=1, values_only=True):
    print(row)  # Chaque 'row' est un tuple de valeurs
