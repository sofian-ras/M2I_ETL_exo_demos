from openpyxl import Workbook

print("=== Démo OpenPyXL - Formules ===")

wb = Workbook()
ws = wb.active

# En-têtes
ws['A1'] = 'Quantité'
ws['B1'] = 'Prix unitaire'
ws['C1'] = 'Total'

# Données
ws['A2'] = 5
ws['B2'] = 10.50

# Formule dans Excel : C2 = A2 * B2
ws['C2'] = '=A2*B2'

# Autres exemples de formules
ws['A10'] = '=SUM(A2:A9)'      # Somme des quantités A2 à A9
ws['B10'] = '=AVERAGE(B2:B9)'  # Moyenne des prix unitaires
ws['C10'] = '=MAX(C2:C9)'      # Maximum des totaux

# Important : openpyxl n'évalue pas les formules.
# Les cellules de type "=..." seront calculées lorsqu'on ouvrira le fichier dans Excel.

wb.save('with_formulas.xlsx')
print("Fichier 'with_formulas.xlsx' créé (formules à calculer dans Excel).")
