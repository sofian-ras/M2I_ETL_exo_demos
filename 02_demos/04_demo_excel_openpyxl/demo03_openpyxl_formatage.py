from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("=== Démo OpenPyXL - Formatage ===")

wb = Workbook()
ws = wb.active

# Écrire une valeur dans la cellule A1
ws['A1'] = 'Titre'

# Police (font) : Arial, taille 14, gras, couleur rouge (hex FF0000)
ws['A1'].font = Font(
    name='Arial',
    size=14,
    bold=True,
    color='FF0000'
)

# Couleur de fond (fond jaune)
ws['A1'].fill = PatternFill(
    start_color='FFFF00',
    end_color='FFFF00',
    fill_type='solid'
)

# Alignement horizontal/vertical au centre
ws['A1'].alignment = Alignment(
    horizontal='center',
    vertical='center'
)

# Définition d'une bordure fine pour tous les côtés
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
ws['A1'].border = thin_border

# Largeur de la colonne A
ws.column_dimensions['A'].width = 20

# Facultatif : ajuster aussi la hauteur de la ligne 1
ws.row_dimensions[1].height = 25

wb.save('formatted.xlsx')
print("Fichier 'formatted.xlsx' créé avec formatage sur A1.")
